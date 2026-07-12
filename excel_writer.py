import pandas as pd
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def _write_df(ws, df):
    if df is None or df.empty:
        ws.append(["데이터 없음"])
        return

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)


def _style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            value = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(value))

        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)


def _style_score_sheet(ws):
    # 추천점수 시트 가독성 개선
    for row in range(2, ws.max_row + 1):
        grade = ws.cell(row=row, column=4).value  # 등급 컬럼

        if grade == "★★★★★":
            fill = PatternFill("solid", fgColor="F4CCCC")
        elif grade == "★★★★☆":
            fill = PatternFill("solid", fgColor="FCE5CD")
        elif grade == "★★★☆☆":
            fill = PatternFill("solid", fgColor="FFF2CC")
        else:
            fill = PatternFill("solid", fgColor="FFFFFF")

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    # 주요 컬럼 폭 고정
    widths = {
        "A": 10,   # 점수순위
        "B": 12,   # 종목코드
        "C": 18,   # 종목명
        "D": 10,   # 총점
        "E": 12,   # 등급
        "F": 12,   # 최종추천
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # AI추천사유 컬럼 자동 탐색 후 넓게
    for cell in ws[1]:
        if cell.value == "AI추천사유":
            col_letter = get_column_letter(cell.column)
            ws.column_dimensions[col_letter].width = 70

        if cell.value == "뉴스요약":
            col_letter = get_column_letter(cell.column)
            ws.column_dimensions[col_letter].width = 60


def save_to_excel(
    portfolio_df=None,
    volume_rank_df=None,
    rise_rank_df=None,
    trade_value_df=None,
    signal_df=None,
    scored_df=None,
    chart_history_df=None
):
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"stock_analysis_{now}.xlsx"

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    sheets = [
        ("추천점수", scored_df),
        ("보유종목", portfolio_df),
        ("신호", signal_df),
        ("거래량TOP30", volume_rank_df),
        ("상승률TOP30", rise_rank_df),
        ("거래대금TOP30", trade_value_df),
        ("차트원본", chart_history_df),
    ]

    for sheet_name, df in sheets:
        ws = wb.create_sheet(sheet_name)
        _write_df(ws, df)
        _style_sheet(ws)

        if sheet_name == "추천점수":
            _style_score_sheet(ws)

    # 차트원본은 계산근거용이라 숨김 처리
    if "차트원본" in wb.sheetnames:
        wb["차트원본"].sheet_state = "hidden"

    wb.save(filename)
    print(f"엑셀 저장 완료: {filename}")